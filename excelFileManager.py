# Excel file manager for workscope exhibit generator tool
#
# NOTES: 
#   1. This module was written to run under Python 2.7.x
#   2. This module relies upon the OpenPyXl library being installed
#       OpenPyXl is used to read and navigate the input .xlsx workbook
#   3. To install OpenPyXl under Python 2.7.x:
#       <Python_installation_folder>/python.exe -m pip install openpyxl
#
# The code in this module previously resided in 'workscope_exhibit_tool.py'
#
# Author: Benjamin Krepp
# Date: 8 August 2018
#
# Requirements on the input .xlsx spreadsheet
# ===========================================
#
# This module relies upon the workscope exhibit template .xlsx file having been created
# with 'defined_names' for various locations of interest in it.
# Although 'defined names' are often used to specify ranges of cells, here they are used
# to identify specific cells, whose row and/or column index is obtained for use in the
# generation of the HTML for a workscope exhibit.
#
# 1. The worksheet containing the workscope exhibits MUST be named 'workscope_exhibits'.
#    Other worksheets may be present; their contents are ignored by this script.
#
# 2. These defined names MUST be present in the workbook and defined as follows:
#
#   project_name_cell - cell containing the project's name
#   direct_salary_cell - cell containing the total direct salary and overhead
#   odc_cell - cell containing the total of other direct costs
#   total_cost_cell - cell containing the total cost of the project
#   task_list_top - any cell in line immediately preceeding list of tasks;
#                   only the row index of this cell is used
#   task_list_bottom - any cell in line immediately following list of tasks;
#                      only the row index of this cell is used
#   funding list_top - any cell in line immediately preceeding list of funding
#                      sources; only the row index of this cell is used
#   funding list_bottom - any cell in line immediately following list of funding
#                         sources; only the row index of this cell is used
#   task_number_column - any cell in column containing the task numbers in
#                        the cost table; only the column index of this cell is used
#   task_name_column - any cell in column containint the task name in the 
#                      cost table; only the column index of this cell is used
#   m1_column -    cost table header cell containing the text 'M-1'
#   p5_column -    cost table header cell containing the text 'P-5'
#   p4_column -    cost table header cell containing the text 'P-4' 
#   p3_column -    cost table header cell containing the text 'P-3'
#   p2_column -    cost table header cell containing the text 'P-2'
#   p1_column -    cost table header cell containing the text 'P-1'
#   sp3_column -   cost table header cell containing the text 'SP-3'
#   sp1_column -   cost table header cell containing the text 'SP-1'
#   temp_column -  cost table header cell containing the text 'Temp'
#   total_column - cost table header cell containing the text 'Total'
#   direct_salary_column - cost table header cell containg the string
#                          'Salary' (2nd line of 'Direct Salary')
#   overhead_column - cost table header cell containing the overhead
#                     rate (as a string); used when we want to access the column
#   overhead_cell -   ditto; used when we want to access the cell itself
#   total_cost_column - cost table header cell containing the text
#                       'Cost' (2nd line of 'Total Cost'
#   total_line - any cell in row containing person-hour totals in 
#                the cost table; only the row index of this cell is used
#   odc_travel_line - any cell on line containing Other Direct Costs: travel;
#                     only the row index of this cell is used
#   odc_office_equipment_line - any cell on line containing Other Direct Costs:
#                               general office equipment; only the row index of 
#                               this cell is used
#   odc_dp_equipment_line - any cell on line containing Other Direct Costs:
#                           data processing equipment; only the row index of 
#                           this cell is used
#   odc_consultants_line - any cell on line containing Other Direct Costs:
#                          consultants; only the row index of this cell is used
#   odc_printing_line - any cell on line containing Other Direct Costs:
#                       printing; only the row index of this cell is used
#   odc_other_line - any cell on line containing Ohter Direct Costs: other;
#                    only the row index of this cell is used
#
# Internals of this Module: Top-level Functions
# =============================================
#
# initExcelFile - Reads a completed .xlsx workscope exhibit template; extracts the row-
#                 and colum-indices (and a couple of other things) of interest/use, 
#                 which are stored in a dictionary object. This object is subsequently
#                 used throughout the 'workscope_exhibit_tool.py' module; it is the most 
#                 important data structure in the program as a whole.
#
# get_sched_col_info - Get info about columnar organization of schedule table;
#                      this may be 'merged' into initExcelFile.
#
# Internals of this Module: Utility Functions
# ===========================================
#
# get_column_index - return the column index for a defined name assigned to a single cell
#
# get_row_index - return the row index for a defined name assigned to a single cell
#
# get_cell_contents - returns the contents of a cell, given a worksheet name, row index,
#                     and column index
#
#
# A Guide for the Perplexed (with apologies to Maimonides):
# An 'Ultra-quck' Quick Start Guide to Using the OpenPyXl Library
# ===============================================================
#
# Open an .xlsx workbook:
#   wb = openpyxl.load_workbook(full_path_to_workbook_file, data_only=True)
#
# Get list of worksheets in workbook:
#   ws_list = wb.sheetnames
#
# Get a named worksheet:
#   ws = wb['workscope_template']
#
# Get list of defined names in workbook:
#   list_of_dns = wb.defined_names
#
# Get value of a defined name, e.g., 'foobar'
#   dn_val = wb.defined_names['foobar'].value
#
# Get the worksheet and cell indices for a defined name,
# and get the value of the cell it refers to
#   temp = dn.split('!')
#   # temp[0] is the worksheet name; temp[1] is the cell reference
#   cell = ws[temp[1]]
#   row_index = cell.row
#   column_index = cell.col_idx
#   cell_value = ws.cell(row_index,column_index).value
#
# Get a cell, given a worksheet, and row and column indices
#    cell = ws.cell(163, 24)
#
# Get the fill and fill pattern type of a cell
#    fill = cell.fill
#    patternType = fill.patternType
#
# N.B. The 'magic' fill patternType indicating a filled-in cell in the 
#      schedule exhibit is 'gray125'
#
###############################################################################

import openpyxl

# Return the column index for a defined name assigned to A SINGLE CELL.
# Note: In Excel, the scope of 'defined names' is the entire workBOOK, not a particular workSHEET.
def get_column_index(wb, name):
    ws = wb['workscope_exhibits']
    x = wb.defined_names[name].value
    temp = x.split('!')
    # temp[0] is the worksheet reference, temp[1] is the cell reference
    cell = ws[temp[1]]
    col_ix = cell.col_idx
    return col_ix
# end_def get_column_index()

# Return the row index for a defined name assigned to A SINGLE CELL.
# Note: In Excel, the scope of 'defined names' is the workBOOK, not a particular workSHEET.
def get_row_index(wb, name):
    ws = wb['workscope_exhibits']
    x = wb.defined_names[name].value
    temp = x.split('!')
    # temp[0] is the worksheet reference, temp[1] is the cell reference
    cell = ws[temp[1]]
    row = cell.row
    return row
# end_def get_row_index()

# Return the contents of a cell.
# If OpenPyXl cell accessor raises exception OR value returned by OpenPyXl accessor is None, return the empty string.
def get_cell_contents(ws, row_ix, col_ix):
    try:
        temp = ws.cell(row_ix, col_ix).value
    except:
        temp = ''
    if temp == None:
        retval = ' '
    else:
        retval = temp
    return retval
# end_def get_cell_contents()



# Open the workbook (.xlsx file) inidicated by the "fullpath" parameter.
# Return a dictionary containing all row and column inidices of interest,
# as well as entries for the workbook itself and the worksheet containing
# the workscope exhibit data.
# 
def initExcelFile(fullpath):
    # retval dictionary
    retval = {}
    # Workbook MUST be opened with data_only parameter set to True.
    # This ensures that we read the computed value in cells containing a formula, not the formula itself.
    wb = openpyxl.load_workbook(fullpath, data_only=True)
    retval['wb'] = wb
    # 
    # N.B. The worksheet containing the workscope exhibits is named 'workscope_exhibits'.
    ws = wb['workscope_exhibits']
    retval['ws'] = ws
    
    # Collect row and column indices for cells of interest for Exhibit 2
    #
    try:
        retval['project_name_cell_row_ix'] = get_row_index(wb, 'project_name_cell')
        retval['project_name_cell_col_ix'] = get_column_index(wb, 'project_name_cell')
    except:
        retval['project_name_cell_row_ix'] = None
        retval['project_name_cell_col_ix'] = None
    try:
        retval['direct_salary_cell_row_ix'] = get_row_index(wb, 'direct_salary_cell')
        retval['direct_salary_cell_col_ix'] = get_column_index(wb, 'direct_salary_cell')
    except:
        retval['direct_salary_cell_row_ix'] = None
        retval['direct_salary_cell_col_ix'] = None
    try:
        retval['odc_cell_row_ix'] = get_row_index(wb, 'odc_cell')
        retval['odc_cell_col_ix'] = get_column_index(wb, 'odc_cell')
    except:
        retval['odc_cell_row_ix'] = None
        retval['odc_cell_col_ix'] = None
    try:
        retval['total_cost_cell_row_ix'] = get_row_index(wb, 'total_cost_cell')
        retval['total_cost_cell_col_ix'] = get_column_index(wb, 'total_cost_cell')
    except:
        retval['total_cost_cell_row_ix'] = None
        retval['total_cost_cell_col_ix'] = None
    # Overhead rate cell.
    try:
        retval['overhead_cell_row_ix'] = get_row_index(wb, 'overhead_cell')
        retval['overhead_cell_col_ix'] = get_column_index(wb, 'overhead_cell')
    except:
        retval['overhead_cell_row_ix'] = None
        retval['overhead_cell_col_ix'] = None
    #       
    # Collect useful row indices for Exhibit 2
    #
    try:
        retval['task_list_top_row_ix'] = get_row_index(wb, 'task_list_top')
    except:
        retval['task_list_top_row_ix'] = None
    try:
        retval['task_list_bottom_row_ix'] = get_row_index(wb, 'task_list_bottom')
    except:
        retval['task_list_bottom_row_ix'] = None
    try:
        retval['total_line_row_ix'] = get_row_index(wb, 'total_line')   
    except:
        retval['total_line_row_ix'] = None
    # Rows containing other direct costs
    try:
        retval['odc_travel_line_ix'] =  get_row_index(wb, 'odc_travel_line')
    except:
        retval['odc_travel_line_ix'] = None
    try:
        retval['odc_office_equipment_line_ix'] = get_row_index(wb, 'odc_office_equipment_line')
    except:
        retval['odc_office_equipment_line_ix'] = None
    try:
        retval['odc_dp_equipment_line_ix'] = get_row_index(wb, 'odc_dp_equipment_line')
    except:
        retval['odc_dp_equipment_line_ix'] = None
    try:
        retval['odc_consultants_line_ix'] = get_row_index(wb, 'odc_consultants_line')
    except:
        retval['odc_consultants_line_ix'] = None
    try:
        retval['odc_printing_line_ix'] = get_row_index(wb, 'odc_printing_line')
    except:
        retval['odc_printing_line_ix'] = None
    try:
        retval['odc_other_line_ix'] = get_row_index(wb, 'odc_other_line')   
    except:
        retval['odc_other_line_ix'] =  None
    # Rows containing info on funding source(s)
    try:
        retval['funding_list_top_row_ix'] = get_row_index(wb, 'funding_list_top')
    except:
        retval['funding_list_top_row_ix'] = None
    try:
        retval['funding_list_bottom_row_ix'] = get_row_index(wb, 'funding_list_bottom')
    except:
        retval['funding_list_bottom_row_ix'] = None
    #
    # Collect useful column indices for Exhibit 2
    #
    try:
        retval['task_number_col_ix'] = get_column_index(wb, 'task_number_column')
    except:
        retval['task_number_col_ix'] = None
    try:
        retval['task_name_col_ix'] = get_column_index(wb, 'task_name_column')
    except:
        retval['task_name_col_ix'] = None       
    try:
        retval['m1_col_ix'] = get_column_index(wb, 'm1_column')
    except:
        retval['m1_col_ix'] = None
    try:
        retval['p5_col_ix'] = get_column_index(wb, 'p5_column')
    except:
        retval['p5_col_ix'] = None
    try:    
        retval['p4_col_ix'] = get_column_index(wb, 'p4_column')
    except:
        retval['p4_col_ix'] = None
    try:
        retval['p3_col_ix'] = get_column_index(wb, 'p3_column')
    except:
        retval['p3_col_ix'] = None
    try:
        retval['p2_col_ix'] = get_column_index(wb, 'p2_column')
    except:
        retval['p2_col_ix'] = None
    try:
        retval['p1_col_ix'] = get_column_index(wb, 'p1_column')
    except:
        retval['p1_col_ix'] = None
    try:
        retval['sp3_col_ix'] = get_column_index(wb, 'sp3_column')
    except:
        retval['sp3_col_ix'] = None
    try:
        retval['sp1_col_ix'] = get_column_index(wb, 'sp1_column')
    except:
        retval['sp1_col_ix'] = None
    try:
        retval['temp_col_ix'] = get_column_index(wb, 'temp_column')
    except:
        retval['temp_col_ix'] = None
    # The following statement refers to the column for total labor cost before overhead
    try:
        retval['total_col_ix'] = get_column_index(wb, 'total_column')
    except:
        retval['total_col_ix'] = None
    try:
        retval['direct_salary_col_ix'] = get_column_index(wb, 'direct_salary_column')
    except:
        retval['direct_salary_col_ix'] = None
    try:
        retval['overhead_col_ix'] = get_column_index(wb, 'overhead_column')
    except:
        retval['total_col_ix'] = None
    try:
        retval['total_cost_col_ix'] = get_column_index(wb, 'total_cost_column')
    except:
        retval['total_cost_col_ix'] = None
    #
    # C'est un petit hacque: The column index for funding source names is the same as that for task names.
    #
    retval['funding_source_name_col_ix'] = retval['task_name_col_ix']
    
    # Collect row and column indices for cells of interest for Exhibit 1
    #
    try:
        retval['first_schedule_col_ix'] = get_column_index(wb, 'first_schedule_column')
    except:
        retval['first_schedule_col_ix'] = None
    try:
        retval['last_schedule_col_ix'] = get_column_index(wb, 'last_schedule_column')
    except:
        retval['last_schedule_col_ix'] = None
    try:
        retval['milestone_label_col_ix'] = get_column_index(wb, 'milestone_label_column')
    except:
        retval['milestone_label_col_ix'] = None
    try:
        retval['milestone_name_col_ix'] = get_column_index(wb, 'milestone_name_column')
    except:
        retval['milestone_name_col_ix'] = None
    try:
        retval['milestones_list_first_row_ix'] = get_row_index(wb, 'milestones_list_first_row')
    except:
        retval['milestones_list_first_row_ix'] = None
    #
    # N.B. The last row of the milestones list is found programmatically by crawling down
    #      milestone_label_column until the first row containing a blank cell is found.
        
    return retval
# end_def initExcelFile()

# Collect and compute information on columnar organization
# of the schedule chart in Exhibit 1
def get_sched_col_info(xlsInfo):
    rv = {}
    
    # ***  How to get 'last_week'? Harvest from .xlsx file????
    # Placeholder, for now.
    rv['last_week'] = 48
    
    # Placeholder, for now.
    # *** TBD: Harvest this value from input .xlsx and store in xlsInfo in initialize()
    rv['sched_unit'] = 'w' 
    
    # Comments from CFML code:
    # if the project will span 12 weeks or less, or 6-12 months, the column width will be such that
	# 12 columns would fit in the space alloted (wide columns), other so that 24 columns would fit (narrow)
    
    if rv['last_week'] <= 13:
        rv['sched_col_width_basis'] = 12
        rv['column_unit'] = 'w'
    elif rv['last_week'] <= 25:
        if rv['schedule_unit'] == 'months':
            rv['schedule_col_width_basis'] = 12
            rv['column_unit'] = 'm'
        else:
            rv['schedule_col_width_basis'] = 24
            rv['column_unit'] = 'w'
        # end_if
    elif rv['last_week'] <= 53:
        # Project is one year or less, but moe than six months
        rv['sched_col_width_basis'] = 12
        rv['column_unit'] = 'm'
    elif rv['last_week'] <= 105:
        # Project is two years or less, but more than one year
        rv['sched_col_width_basis'] = 24
        rv['column_unit'] = 'm'
    elif rv['last_week'] < 157:
        # Project is three years or less, but more than two years
        rv['sched_col_width_basis'] = 12
        rv['column_unit'] = 'q'
    else:
        # Project is more than three years long
        # Note, however, that projects more than four years long just won't fit
        rv['sched_col_width_basis'] = 24
        rv['column_unit'] = 'q'
    # end_if
    
    # Each bar unit is a column, and if the column headings are in months, 
    # then each bar unit is slightly over 4 weeks
    if rv['column_unit'] == 'm':
        rv['weeks_per_bar_unit'] = WEEKS_PER_MONTH
    elif column_unit == 'q':
        rv['weeks_per_bar_unit'] = WEEKS_PER_QUARTER
    else:
        rv['weeks_per_bar_unit'] = 1
    # end_if
    
    # Set 'schedule_columns'
    # CFML statment: schedule_columns = Ceiling(Round(100*(last_week - 1)/weeks_per_bar_unit)/100)
    rv['schedule_columns'] = int(math.ceil(round(100*(rv['last_week'] - 1)/rv['weeks_per_bar_unit'],0)/100))
    
    # Debug:
    print 'schedule_columns = ' + str(rv['schedule_columns'])

    # This is now hardwired
    rv['cell_border_unit'] = 'PixBorder'    
    
    return rv
# end_def get_sched_col_info()